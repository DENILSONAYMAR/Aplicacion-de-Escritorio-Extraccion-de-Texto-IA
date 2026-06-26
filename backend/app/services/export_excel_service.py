import json
import os
import re
from datetime import datetime
from datetime import time as datetime_time
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app import models


INTERNAL_CONTEXT_MARKER = "--- CONTEXTO INTERNO DEL SISTEMA ---"

HEADER_FILL = PatternFill(start_color="D9EAF7",
                          end_color="D9EAF7", fill_type="solid")
OK_FILL = PatternFill(start_color="FFFFFF",
                      end_color="FFFFFF", fill_type="solid")
ILLEGIBLE_FILL = PatternFill(
    start_color="FECACA", end_color="FECACA", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2EC"),
    right=Side(style="thin", color="D9E2EC"),
    top=Side(style="thin", color="D9E2EC"),
    bottom=Side(style="thin", color="D9E2EC"),
)

CONFIDENCE_SCORE = {
    "alta": 4,
    "media": 3,
    "baja": 2,
    "ninguna": 1,
    None: 0,
}

STATUS_SCORE = {
    "ok": 5,
    "inferido": 4,
    "requiere_revision": 3,
    "pending_review": 2,
    "campo_no_encontrado": 1,
    "ilegible": 0,
    "error_formato": 0,
    "rechazado": 0,
    None: 0,
}


def _safe_filename(text: str) -> str:
    text = text or "extraccion"
    invalid = ['\\', '/', ':', '*', '?', '"', '<', '>', '|']

    for char in invalid:
        text = text.replace(char, "_")

    return text.strip()[:80] or "extraccion"


def _visible_instruction_only(text: str | None, fallback: str = "") -> str:
    text = (text or "").strip()

    if not text:
        return fallback

    if INTERNAL_CONTEXT_MARKER in text:
        text = text.split(INTERNAL_CONTEXT_MARKER)[0].strip()

    return text or fallback


def _parse_visible_payload(text: str | None, fallback: str = "") -> dict:
    visible = _visible_instruction_only(text, fallback=fallback)

    try:
        data = json.loads(visible)

        if isinstance(data, dict):
            return {
                "encabezado": data.get("encabezado") or fallback,
                "contexto": data.get("contexto") or "",
                "tipo_resultado": data.get("tipo_resultado") or "",
                "validacion": data.get("validacion") or "",
            }
    except Exception:
        pass

    return {
        "encabezado": fallback,
        "contexto": visible if visible != fallback else "",
        "tipo_resultado": "",
        "validacion": "",
    }


def _clean_value(value) -> str:
    if value is None:
        return "no visible"

    value = str(value).strip()

    if not value:
        return "no visible"

    if value.lower() in ["null", "none", "nan", "[campo no encontrado]"]:
        return "no visible"

    if value.lower() in ["[ilegible]"]:
        return "ilegible"

    return value


def _result_type_key(label: str | None) -> str:
    text = (label or "").strip().lower()

    aliases = {
        "fecha": "date",
        "hora": "time",
        "texto": "text",
        "decimal": "decimal",
        "numero": "integer",
        "número": "integer",
        "calculo": "calculation",
        "cálculo": "calculation",
    }

    return aliases.get(text, text)


def _field_type_from_description(field: models.ExtractionTemplateField) -> str:
    payload = _parse_visible_payload(field.description, fallback=field.field_name)
    return _result_type_key(payload.get("tipo_resultado")) or "text"


def _parse_decimal(value: str):
    text = value.strip()

    if not re.search(r"\d", text):
        return None

    text = re.sub(r"[^\d,.\-]", "", text)

    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")

    try:
        return float(text)
    except ValueError:
        return None


def _parse_date(value: str):
    text = value.strip()
    formats = [
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%Y-%m-%d",
        "%d/%m/%y",
        "%d-%m-%y",
    ]

    for date_format in formats:
        try:
            return datetime.strptime(text, date_format)
        except ValueError:
            continue

    return None


def _parse_time(value: str):
    match = re.search(r"\b(\d{1,2})[:.](\d{2})(?:\s*([ap]\.?m\.?))?\b", value, re.I)

    if not match:
        return None

    hour = int(match.group(1))
    minute = int(match.group(2))
    meridian = (match.group(3) or "").lower().replace(".", "")

    if meridian == "pm" and hour < 12:
        hour += 12
    elif meridian == "am" and hour == 12:
        hour = 0

    if hour > 23 or minute > 59:
        return None

    return datetime_time(hour=hour, minute=minute)


def _parse_boolean(value: str):
    text = value.strip().lower()

    yes_values = ["si", "sí", "true", "verdadero", "correcto", "ok", "cumple", "aprobado"]
    no_values = ["no", "false", "falso", "incorrecto", "no cumple", "rechazado"]

    if text in yes_values:
        return True

    if text in no_values:
        return False

    return None


def _excel_value_and_format(value: str, result_type: str):
    if value.lower() in ["no aplica", "no visible", "ilegible"]:
        return value, None

    if result_type == "integer":
        parsed = _parse_decimal(value)
        if parsed is not None:
            return int(round(parsed)), "0"

    if result_type == "decimal":
        parsed = _parse_decimal(value)
        if parsed is not None:
            return parsed, "#,##0.00"

    if result_type == "date":
        parsed = _parse_date(value)
        if parsed is not None:
            return parsed, "dd/mm/yyyy"

    if result_type == "time":
        parsed = _parse_time(value)
        if parsed is not None:
            return parsed, "hh:mm"

    if result_type == "calculation":
        parsed = _parse_boolean(value)
        if parsed is not None:
            return parsed, None

    return value, None


def _result_has_value(result: models.ExtractionResult) -> bool:
    value = result.normalized_value or result.raw_value

    if value is None:
        return False

    value = str(value).strip()

    if not value:
        return False

    if value.lower() in ["null", "none", "nan", "no visible", "ilegible", "no aplica"]:
        return False

    return True


def _result_score(result: models.ExtractionResult) -> int:
    score = 0

    if _result_has_value(result):
        score += 20

    if not result.needs_review:
        score += 10

    score += STATUS_SCORE.get(result.status, 0)
    score += CONFIDENCE_SCORE.get(result.confidence_level, 0)

    return score


def _get_best_result_for_field(
    document_results: list[models.ExtractionResult],
    field_name: str,
) -> models.ExtractionResult | None:
    candidates = [
        result
        for result in document_results
        if result.field_name == field_name
    ]

    if not candidates:
        return None

    candidates.sort(key=_result_score, reverse=True)

    return candidates[0]


def _get_final_value(result: models.ExtractionResult | None) -> str:
    if not result:
        return "no aplica"

    value = result.normalized_value or result.raw_value

    return _clean_value(value)


def _get_review_filter_field():
    if not hasattr(models, "HumanReview"):
        return None

    if hasattr(models.HumanReview, "extraction_result_id"):
        return models.HumanReview.extraction_result_id

    if hasattr(models.HumanReview, "result_id"):
        return models.HumanReview.result_id

    return None


def _get_latest_review(
    db: Session,
    result: models.ExtractionResult | None,
):
    if not result:
        return None

    if not hasattr(models, "HumanReview"):
        return None

    filter_field = _get_review_filter_field()

    if filter_field is None:
        return None

    query = (
        db.query(models.HumanReview)
        .filter(filter_field == result.id)
    )

    if hasattr(models.HumanReview, "id"):
        query = query.order_by(models.HumanReview.id.desc())

    return query.first()


def _get_review_status(review) -> str:
    if not review:
        return ""

    return str(getattr(review, "review_status", "") or "").strip().lower()


def _get_review_note(review) -> str:
    if not review:
        return ""

    return str(getattr(review, "reviewer_notes", "") or "").strip()


def _was_marked_illegible(review) -> bool:
    return _get_review_status(review) == "marked_illegible"


def _was_accepted_or_corrected(review) -> bool:
    return _get_review_status(review) in ["accepted", "corrected"]


def _is_valid_result(
    result: models.ExtractionResult | None,
    review=None,
) -> bool:
    if not result:
        return False

    if review and _was_marked_illegible(review):
        return False

    if review and _was_accepted_or_corrected(review):
        return True

    value = _get_final_value(result).lower()

    return (
        value not in ["no aplica", "no visible", "ilegible"]
        and result.status == "ok"
        and result.needs_review is False
        and result.confidence_level == "alta"
        and result.source_type != "no_visible"
    )


def _style_header(cell):
    cell.fill = HEADER_FILL
    cell.font = Font(bold=True)
    cell.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER


def _style_body(cell):
    cell.fill = OK_FILL
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = THIN_BORDER


def _autosize_columns(sheet, max_width: int = 42):
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = cell.value

            if value is None:
                continue

            max_length = max(max_length, len(str(value)))

        sheet.column_dimensions[column_letter].width = min(
            max(max_length + 2, 12), max_width)


def _get_template_fields(
    db: Session,
    template_id: int | None,
    results: list[models.ExtractionResult],
) -> list[models.ExtractionTemplateField]:
    if template_id is not None:
        template_fields = (
            db.query(models.ExtractionTemplateField)
            .filter(models.ExtractionTemplateField.template_id == template_id)
            .order_by(models.ExtractionTemplateField.id.asc())
            .all()
        )

        if template_fields:
            return template_fields

    field_names = sorted({
        result.field_name
        for result in results
        if result.field_name
    })

    return [
        models.ExtractionTemplateField(
            field_name=field_name,
            display_name=field_name,
            description=field_name,
            target_location=None,
            required=False,
        )
        for field_name in field_names
    ]


def _group_results_by_document_instance(
    results: list[models.ExtractionResult],
) -> dict[str, list[models.ExtractionResult]]:
    grouped = defaultdict(list)

    for result in results:
        key = f"{result.job_id}::{result.document_id or result.file_name or result.id}"
        grouped[key].append(result)

    return dict(grouped)


def _write_summary_sheet(
    db: Session,
    workbook: Workbook,
    results: list[models.ExtractionResult],
    template_fields: list[models.ExtractionTemplateField],
):
    sheet = workbook.active
    sheet.title = "Resumen"

    field_names = [field.field_name for field in template_fields]
    field_types = {
        field.field_name: _field_type_from_description(field)
        for field in template_fields
    }
    headers = ["ID"] + field_names + ["Observaciones"]

    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_index, value=header)
        _style_header(cell)

    grouped = _group_results_by_document_instance(results)

    row_index = 2
    sequential_id = 1

    for _, document_results in grouped.items():
        id_cell = sheet.cell(row=row_index, column=1, value=sequential_id)
        _style_body(id_cell)

        observations = []

        for field_offset, field_name in enumerate(field_names, start=2):
            best_result = _get_best_result_for_field(
                document_results=document_results,
                field_name=field_name,
            )

            review = _get_latest_review(db, best_result)
            value = _get_final_value(best_result)
            excel_value, number_format = _excel_value_and_format(
                value,
                field_types.get(field_name, "text"),
            )

            note = _get_review_note(review)

            if note:
                observations.append(f"{field_name}: {note}")

            cell = sheet.cell(row=row_index, column=field_offset, value=excel_value)
            _style_body(cell)

            if number_format:
                cell.number_format = number_format

        observation_cell = sheet.cell(
            row=row_index,
            column=len(headers),
            value="\n".join(observations) if observations else "",
        )

        _style_body(observation_cell)

        row_index += 1
        sequential_id += 1

    sheet.freeze_panes = "A2"
    _autosize_columns(sheet, max_width=50)


def _field_validation_percentage(
    db: Session,
    field_name: str,
    grouped_results: dict[str, list[models.ExtractionResult]],
) -> str:
    total_documents = len(grouped_results)

    if total_documents == 0:
        return "0%"

    valid_count = 0

    for _, document_results in grouped_results.items():
        best_result = _get_best_result_for_field(
            document_results=document_results,
            field_name=field_name,
        )

        review = _get_latest_review(db, best_result)

        if _is_valid_result(best_result, review):
            valid_count += 1

    percentage = round((valid_count / total_documents) * 100, 2)

    if percentage.is_integer():
        return f"{int(percentage)}%"

    return f"{percentage}%"


def _write_instructions_sheet(
    db: Session,
    workbook: Workbook,
    template_fields: list[models.ExtractionTemplateField],
    grouped_results: dict[str, list[models.ExtractionResult]],
):
    sheet = workbook.create_sheet(title="Instrucciones para encabezados")

    headers = [
        "ID",
        "Encabezados",
        "Contexto",
        "Porcentaje de extracción validado",
    ]

    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_index, value=header)
        _style_header(cell)

    for row_index, field in enumerate(template_fields, start=2):
        payload = _parse_visible_payload(
            field.description,
            fallback=field.field_name,
        )

        # Solo guardamos el contexto escrito por el usuario.
        # No se mezcla tipo de dato ni reglas internas.
        context = payload.get("contexto") or ""

        values = [
            row_index - 1,
            payload.get("encabezado") or field.field_name,
            context,
            _field_validation_percentage(
                db=db,
                field_name=field.field_name,
                grouped_results=grouped_results,
            ),
        ]

        for col_index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=col_index, value=value)
            _style_body(cell)

    sheet.freeze_panes = "A2"
    _autosize_columns(sheet, max_width=80)


def _write_internal_audit_sheet(
    workbook: Workbook,
    results: list[models.ExtractionResult],
):
    sheet = workbook.create_sheet(title="_auditoria")
    sheet.sheet_state = "hidden"

    headers = [
        "result_id",
        "job_id",
        "document_id",
        "file_name",
        "page_number",
        "field_name",
        "raw_value",
        "normalized_value",
        "status",
        "confidence_level",
        "needs_review",
        "source_type",
        "evidence_text",
    ]

    for col_index, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_index, value=header)

    for row_index, result in enumerate(results, start=2):
        values = [
            result.id,
            result.job_id,
            result.document_id,
            result.file_name,
            result.page_number,
            result.field_name,
            result.raw_value,
            result.normalized_value,
            result.status,
            result.confidence_level,
            result.needs_review,
            result.source_type,
            result.evidence_text,
        ]

        for col_index, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=col_index, value=value)


def export_job_results_to_excel(
    db: Session,
    job_id: int,
    template_id: int | None = None,
) -> models.ExportFile:
    job = (
        db.query(models.ExtractionJob)
        .filter(models.ExtractionJob.id == job_id)
        .first()
    )

    if not job:
        raise ValueError("Job de extracción no encontrado")

    project = (
        db.query(models.ExtractionProject)
        .filter(models.ExtractionProject.id == job.project_id)
        .first()
    )

    if not project:
        raise ValueError("Proyecto de extracción no encontrado")

    results = (
        db.query(models.ExtractionResult)
        .filter(models.ExtractionResult.job_id == job_id)
        .order_by(
            models.ExtractionResult.file_name.asc(),
            models.ExtractionResult.page_number.asc(),
            models.ExtractionResult.id.asc(),
        )
        .all()
    )

    if not results:
        raise ValueError("El job no tiene resultados para exportar")

    template_fields = _get_template_fields(
        db=db,
        template_id=template_id,
        results=results,
    )

    if not template_fields:
        raise ValueError("No hay encabezados disponibles para exportar")

    grouped_results = _group_results_by_document_instance(results)

    storage_root = os.getenv("STORAGE_ROOT", "/storage")
    output_folder = os.path.join(storage_root, "exports")
    os.makedirs(output_folder, exist_ok=True)

    project_name = _safe_filename(project.name)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    file_name = f"{project_name}_{timestamp}.xlsx"
    file_path = os.path.join(output_folder, file_name)

    workbook = Workbook()

    _write_summary_sheet(
        db=db,
        workbook=workbook,
        results=results,
        template_fields=template_fields,
    )

    _write_instructions_sheet(
        db=db,
        workbook=workbook,
        template_fields=template_fields,
        grouped_results=grouped_results,
    )

    _write_internal_audit_sheet(
        workbook=workbook,
        results=results,
    )

    workbook.save(file_path)

    export_file = models.ExportFile(
        job_id=job_id,
        document_id=None,
        export_type="excel",
        file_path=file_path,
        has_warnings=False,
    )

    db.add(export_file)
    db.commit()
    db.refresh(export_file)

    return export_file
