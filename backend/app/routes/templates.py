import io
import re

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from openpyxl import load_workbook

from app.db import get_db
from app import models, schemas

router = APIRouter(prefix="/templates", tags=["templates"])


ITEM_PATTERN = re.compile(r"^\s*item\s*0*([0-9]+)\s*$", re.IGNORECASE)


def _cell_to_text(value) -> str:
    if value is None:
        return ""

    return str(value).strip()


def _parse_option_with_context(text: str) -> dict:
    clean = _cell_to_text(text)

    if not clean:
        return {
            "raw": "",
            "value": "",
            "context": None,
        }

    match = re.match(r"^(.*?)\s*\((.*?)\)\s*$", clean)

    if match:
        value = match.group(1).strip()
        context = match.group(2).strip()

        return {
            "raw": clean,
            "value": value,
            "context": context or None,
        }

    return {
        "raw": clean,
        "value": clean,
        "context": None,
    }


def _detect_item_cells(sheet):
    item_cells = []

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            value = _cell_to_text(sheet.cell(row=row, column=col).value)
            match = ITEM_PATTERN.match(value)

            if match:
                item_cells.append(
                    {
                        "item_index": int(match.group(1)),
                        "item_name": value,
                        "row": row,
                        "col": col,
                    }
                )

    return item_cells


def _collect_options_for_item(sheet, item_cell, item_cells):
    row = item_cell["row"]
    col = item_cell["col"]

    blocking_rows = [
        other["row"]
        for other in item_cells
        if other["col"] == col and other["row"] > row
    ]

    stop_row = min(blocking_rows) if blocking_rows else sheet.max_row + 1

    options = []

    for current_row in range(row + 1, stop_row):
        value = _cell_to_text(sheet.cell(row=current_row, column=col).value)

        if not value:
            continue

        if ITEM_PATTERN.match(value):
            break

        parsed = _parse_option_with_context(value)

        if parsed["value"]:
            options.append(parsed)

    return options


@router.get("/")
def templates_status():
    return {
        "message": "Módulo de plantillas activo",
        "status": "ok",
    }


@router.post("/parse-validation-excel")
async def parse_validation_excel(
    file: UploadFile = File(...),
):
    """
    Lee una base Excel de validación estricta.

    Regla flexible:
    - ITEM 1, ITEM 2, ITEM 7, etc. pueden estar en cualquier celda.
    - Los valores de cada ITEM se leen hacia abajo en la misma columna.
    - Si aparece otro ITEM debajo en la misma columna, se detiene la lectura del ITEM anterior.
    - Cada valor puede tener contexto entre paréntesis:
      Almuerzo (de 12 pm a 6 pm)
    """

    filename = file.filename or ""

    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(
            status_code=400,
            detail="Solo se aceptan archivos .xlsx o .xlsm",
        )

    try:
        content = await file.read()
        workbook = load_workbook(io.BytesIO(content), data_only=True)
        sheet = workbook.active

        item_cells = _detect_item_cells(sheet)

        if not item_cells:
            raise HTTPException(
                status_code=400,
                detail="No se encontraron ITEMS. Usa textos como ITEM 1, ITEM 2, ITEM 7, etc.",
            )

        items_by_index = {}

        for item_cell in item_cells:
            item_index = item_cell["item_index"]
            options = _collect_options_for_item(sheet, item_cell, item_cells)

            if item_index not in items_by_index:
                items_by_index[item_index] = {
                    "item_index": item_index,
                    "item_name": item_cell["item_name"],
                    "options": [],
                }

            existing_raw = {option["raw"]
                            for option in items_by_index[item_index]["options"]}

            for option in options:
                if option["raw"] not in existing_raw:
                    items_by_index[item_index]["options"].append(option)
                    existing_raw.add(option["raw"])

        items = sorted(items_by_index.values(),
                       key=lambda item: item["item_index"])

        return {
            "file_name": filename,
            "sheet_name": sheet.title,
            "items": items,
        }

    except HTTPException:
        raise

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error leyendo base de validación: {str(e)}",
        )


@router.post("", response_model=schemas.ExtractionTemplateResponse)
def create_template(
    payload: schemas.ExtractionTemplateCreate,
    db: Session = Depends(get_db),
):
    if payload.project_id is not None:
        project = (
            db.query(models.ExtractionProject)
            .filter(models.ExtractionProject.id == payload.project_id)
            .first()
        )

        if not project:
            raise HTTPException(
                status_code=404,
                detail="Proyecto de extracción no encontrado",
            )

    template = models.ExtractionTemplate(
        project_id=payload.project_id,
        name=payload.name,
        file_path=payload.file_path,
        template_type=payload.template_type,
    )

    db.add(template)
    db.commit()
    db.refresh(template)

    for field_payload in payload.fields:
        field = models.ExtractionTemplateField(
            template_id=template.id,
            field_name=field_payload.field_name,
            display_name=field_payload.display_name,
            target_location=field_payload.target_location,
            required=field_payload.required,
            description=field_payload.description,
        )

        db.add(field)

    db.commit()
    db.refresh(template)

    return template


@router.get("", response_model=list[schemas.ExtractionTemplateResponse])
def list_templates(
    db: Session = Depends(get_db),
):
    templates = (
        db.query(models.ExtractionTemplate)
        .join(
            models.ExtractionProject,
            models.ExtractionTemplate.project_id == models.ExtractionProject.id,
            isouter=True,
        )
        .filter(
            (models.ExtractionProject.status != "temporary")
            | (models.ExtractionProject.id.is_(None))
        )
        .order_by(models.ExtractionTemplate.created_at.desc())
        .all()
    )

    return templates


@router.get("/{template_id}", response_model=schemas.ExtractionTemplateResponse)
def get_template(
    template_id: int,
    db: Session = Depends(get_db),
):
    template = (
        db.query(models.ExtractionTemplate)
        .filter(models.ExtractionTemplate.id == template_id)
        .first()
    )

    if not template:
        raise HTTPException(
            status_code=404,
            detail="Plantilla no encontrada",
        )

    return template


@router.put("/{template_id}/fields/replace", response_model=schemas.ExtractionTemplateResponse)
def replace_template_fields(
    template_id: int,
    payload: list[schemas.ExtractionTemplateFieldCreate],
    db: Session = Depends(get_db),
):
    template = (
        db.query(models.ExtractionTemplate)
        .filter(models.ExtractionTemplate.id == template_id)
        .first()
    )

    if not template:
        raise HTTPException(
            status_code=404,
            detail="Plantilla no encontrada",
        )

    old_fields = (
        db.query(models.ExtractionTemplateField)
        .filter(models.ExtractionTemplateField.template_id == template_id)
        .all()
    )

    for field in old_fields:
        db.delete(field)

    db.commit()

    for field_payload in payload:
        field = models.ExtractionTemplateField(
            template_id=template_id,
            field_name=field_payload.field_name,
            display_name=field_payload.display_name,
            target_location=field_payload.target_location,
            required=field_payload.required,
            description=field_payload.description,
        )

        db.add(field)

    db.commit()
    db.refresh(template)

    return template
